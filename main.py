import openpyxl
from os import listdir
from openpyxl.styles import Border, Side, Font,  Alignment
from os.path import isfile, join
import datetime
import itertools

borda_fina = Side(border_style="thin", color="000000")
borda = Border(top=borda_fina, left=borda_fina, right=borda_fina, bottom=borda_fina)

transacoes = openpyxl.load_workbook('lista.xlsx').active

planilhas_nomes_arquivos =  [f for f in listdir('./Planilhas') if isfile(join('./Planilhas', f))]
planilhas_nomes = [nome.replace('.xlsx','') for nome in planilhas_nomes_arquivos]
planilhas = {nome:openpyxl.load_workbook(f"./Planilhas/{nome}.xlsx") for nome in planilhas_nomes}


def gerar_ids_planilha():
    ids ={}
    for nome,planilha in planilhas.items():
        tabela = planilha.active
        for row in tabela.rows:
            id = row[0].value
            ids.setdefault(nome,[]).append(id)
            
    return ids

def atualizar_estilos(tabela):
    ultima_linha = tabela.max_row
    for linha in range(4,ultima_linha+1):
        tabela.cell(row=linha,column=2)
        tabela.cell(row=linha,column=2).alignment = Alignment(horizontal="left", vertical="center")
        tabela.cell(row=linha,column=2).font = Font(size=11,name='Calibri')
        tabela.cell(row=linha,column=2).border = borda
        
        tabela.cell(row=linha,column=5).alignment = Alignment(horizontal="left", vertical="center")
        tabela.cell(row=linha,column=5).font = Font(size=11,name='Calibri')
        tabela.cell(row=linha,column=5).border = borda
        
        tabela.cell(row=linha,column=6).alignment = Alignment(horizontal="left", vertical="center")
        tabela.cell(row=linha,column=6).font = Font(size=11,name='Calibri')
        tabela.cell(row=linha,column=6).border = borda
        
        tabela.cell(row=linha,column=11).number_format = '[$R$-416]* #,##0.00;-[$R$-416] #,##0.00'
        tabela.cell(row=linha,column=11).border = borda
        tabela.cell(row=linha,column=11).font = Font(size=10,name='Arial')
        
        tabela.cell(row=linha,column=13).alignment = Alignment(horizontal="right", vertical="center")
        tabela.cell(row=linha,column=13).font = Font(size=11,name='Calibri')
        tabela.cell(row=linha,column=13).number_format="D-MMM-YY"
        tabela.cell(row=linha,column=13).border = borda

def obter_tabela(descricao,tabelas_nomes):
    for nome in tabelas_nomes:
        if nome.lower() in descricao:
            return nome

def extrair_tipo(descricao,tipos):
    for tipo in tipos:
        if tipo in descricao.split():
            return tipo



def obter_ultima_linha(tabela):
    index = 3
    for linha in itertools.islice(tabela.rows,3,None):
        if not linha[1].value:
            return index
        index = index +1
    return index
def escrever_linha(tabela,id,destino,descricao,valor,data,planilha_nome,tipos):
    
    ultima_linha = obter_ultima_linha(tabela)
    
    tipo = extrair_tipo(descricao,tipos)
    
    tabela.cell(row=ultima_linha+1,column=1,value=id)       
    tabela.cell(row=ultima_linha+1,column=2,value=destino)
    if tipo:
        tabela.cell(row=ultima_linha+1,column=5,value=tipo)
        descricao = descricao.replace(tipo,'')
    descricao = descricao.replace(planilha_nome.lower(),'')
    tabela.cell(row=ultima_linha+1,column=6,value=descricao)
    tabela.cell(row=ultima_linha+1,column=11,value=valor)
    tabela.cell(row=ultima_linha+1,column=13,value=data)
    tabela.merge_cells(start_row=ultima_linha+1,end_row=ultima_linha+1,start_column=2,end_column=4)
    tabela.merge_cells(start_row=ultima_linha+1,end_row=ultima_linha+1,start_column=6,end_column=10)
    tabela.merge_cells(start_row=ultima_linha+1,end_row=ultima_linha+1,start_column=11,end_column=12)
    tabela.merge_cells(start_row=ultima_linha+1,end_row=ultima_linha+1,start_column=13,end_column=14)

def obter_dados_transacoes(planilha):
    dados = []
    for linha in planilha.rows:
        tipo = linha[1].value
        if tipo !='PAGAMENTO':
            continue
        descricao = linha[10].value
        if not descricao:
            continue
        
        destino = linha[2].value
        data = linha[3].value
        
        data = list(data)
        dia = int(''.join(data[0:2]))
        mes = int(''.join(data[3:5]))
        ano = int(''.join(data[6:]))
        data = datetime.date(ano, mes, dia)
            
        valor = linha[5].value
        
        valor = float(linha[5].value.replace("R$","").replace(".","").replace(",","."))
        
        id = linha[6].value
        
        dados.append({
            'id':id,
            'data':data,
            'valor':valor,
            'destino':destino,
            'descricao':descricao.lower()
        })
    dados.reverse()
    return dados
    
    
def obter_tipos(tabela):
    tipos = []
    
    for linha in itertools.islice(tabela.rows,3,None):
        tipo =linha[16].value
        if tipo:
            tipos.append(tipo.lower())
    return tipos

dados = obter_dados_transacoes(transacoes)
planilhas_alteradas = []

planilhas_ids = gerar_ids_planilha()

for dado in dados:
    planilha_nome = obter_tabela(dado['descricao'],planilhas_nomes)
    
    if not planilha_nome:
        continue
    if dado['id'] in planilhas_ids[planilha_nome]:
        continue
    tabela = planilhas[planilha_nome].active
    tipos = obter_tipos(tabela)
    
    escrever_linha(tabela,**dado,planilha_nome=planilha_nome,tipos=tipos)
    if planilha_nome not in planilhas_alteradas:
        planilhas_alteradas.append(planilha_nome)


for planilha_nome in planilhas_alteradas:
    tabela = planilhas[planilha_nome].active
    atualizar_estilos(tabela)

    
for nome,planilha in planilhas.items():
    planilha.save(f"./Planilhas/{nome}.xlsx")